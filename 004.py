#!/usr/bin/env python3

from __future__ import annotations

import os
import re
import shutil
import sys
import time
import traceback
import zlib
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange


SOURCE_DIR_NAME = "003_output"
TARGET_DIR_NAME = "004_output"
DEBUG_FILE_NAME = "debug.txt"
BAR_WIDTH = 32
PART_PATTERN = re.compile(r"^(?P<base>.+)\((?P<index>\d+)\)$")
CLEAN_PATTERN = re.compile(r"[^А-Яа-я0-9]+")


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def part_key(path: Path) -> tuple[str, int]:
    match = PART_PATTERN.match(path.stem)
    if match:
        return match.group("base"), int(match.group("index"))
    return path.stem, 1


def has_second_part(source_paths: list[Path]) -> bool:
    return any(part_key(path)[1] == 2 for path in source_paths)


def collect_jobs(source_dir: Path) -> list[list[Path]]:
    grouped_paths: dict[tuple[Path, str], list[Path]] = {}

    for workbook_path in iter_workbooks(source_dir):
        relative_dir = workbook_path.parent.relative_to(source_dir)
        base_name, _ = part_key(workbook_path)
        grouped_paths.setdefault((relative_dir, base_name), []).append(workbook_path)

    jobs: list[list[Path]] = []
    for _, paths in sorted(grouped_paths.items()):
        paths.sort(key=lambda path: part_key(path)[1])
        jobs.append(paths)

    return jobs


def target_paths_for_group(
    source_paths: list[Path],
    source_dir: Path,
    target_dir: Path,
) -> tuple[Path, Path]:
    first_path = sorted(source_paths, key=lambda path: part_key(path)[1])[0]
    relative_dir = first_path.parent.relative_to(source_dir)
    base_name, _ = part_key(first_path)
    target_workbook_path = target_dir / relative_dir / f"{base_name}.xlsx"
    target_json_path = target_workbook_path.with_suffix(".json")
    return target_workbook_path, target_json_path


def clean_label(value) -> str:
    if value is None:
        return ""
    return CLEAN_PATTERN.sub("", str(value)).lower()


def is_integer_like(value) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped.isdigit()
    return False


def column_signature(worksheet, column_index: int) -> tuple[str, ...]:
    code_row = find_code_row(worksheet)
    row_limit = code_row if code_row is not None else min(worksheet.max_row, 10)
    values: list[str] = []

    for row in range(1, row_limit + 1):
        cleaned = clean_label(worksheet.cell(row=row, column=column_index).value)
        if cleaned:
            values.append(cleaned)

    return tuple(values)


def find_code_row(worksheet) -> int | None:
    scan_width = min(worksheet.max_column, 10)

    for row in range(1, worksheet.max_row + 1):
        values = [worksheet.cell(row=row, column=column).value for column in range(1, scan_width + 1)]
        if len(values) < 2:
            continue

        if not is_integer_like(values[0]) or int(float(values[0])) != 1:
            continue
        if not is_integer_like(values[1]) or int(float(values[1])) != 2:
            continue

        numeric_count = sum(1 for value in values if is_integer_like(value))
        if numeric_count >= 3:
            return row

    return None


def descriptive_crc(worksheet) -> int | None:
    if worksheet.max_column < 3:
        return None

    code_row = find_code_row(worksheet)
    if code_row is None:
        return None

    start_row = code_row + 1
    if worksheet.max_row - start_row + 1 < 3:
        return None

    rows = []
    end_row = min(worksheet.max_row, start_row + 4)
    for row in range(start_row, end_row + 1):
        row_values = []
        for column in range(1, 4):
            value = worksheet.cell(row=row, column=column).value
            row_values.append("" if value is None else str(value))
        rows.append("\t".join(row_values))

    if not rows:
        return None

    return zlib.crc32("\n".join(rows).encode("utf-8"))


def descriptive_columns(base_ws, part_ws) -> int:
    limit = min(7, base_ws.max_column, part_ws.max_column)
    count = 0

    for column in range(1, limit + 1):
        if column_signature(base_ws, column) != column_signature(part_ws, column):
            break
        count += 1

    if count < 1:
        raise ValueError("Could not identify descriptive columns")

    if count >= 3:
        base_crc = descriptive_crc(base_ws)
        part_crc = descriptive_crc(part_ws)
        if base_crc is not None and part_crc is not None and base_crc != part_crc:
            raise ValueError("CRC32 mismatch in descriptive rows")

    return count


def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def copy_column_block(
    target_ws,
    source_ws,
    source_start_col: int,
    source_end_col: int,
    target_start_col: int,
) -> None:
    column_offset = target_start_col - source_start_col

    for source_col in range(source_start_col, source_end_col + 1):
        target_col = source_col + column_offset
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)

        if source_ws.column_dimensions[source_letter].width is not None:
            target_ws.column_dimensions[target_letter].width = source_ws.column_dimensions[source_letter].width

        for row in range(1, source_ws.max_row + 1):
            copy_cell(
                source_ws.cell(row=row, column=source_col),
                target_ws.cell(row=row, column=target_col),
            )

    for row_index, dimension in source_ws.row_dimensions.items():
        if dimension.height is not None and target_ws.row_dimensions[row_index].height is None:
            target_ws.row_dimensions[row_index].height = dimension.height

    for cell_range in source_ws.merged_cells.ranges:
        if cell_range.min_col <= source_start_col - 1:
            continue

        shifted_range = CellRange(
            min_col=cell_range.min_col + column_offset,
            min_row=cell_range.min_row,
            max_col=cell_range.max_col + column_offset,
            max_row=cell_range.max_row,
        )
        target_ws.merge_cells(str(shifted_range))


def merge_group(source_paths: list[Path], source_dir: Path, target_dir: Path) -> int:
    source_paths = sorted(source_paths, key=lambda path: part_key(path)[1])
    first_path = source_paths[0]
    target_workbook_path, target_json_path = target_paths_for_group(source_paths, source_dir, target_dir)

    target_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(first_path, target_workbook_path)

    source_json_path = first_path.with_suffix(".json")
    if source_json_path.exists():
        shutil.copy2(source_json_path, target_json_path)

    if len(source_paths) == 1 or not has_second_part(source_paths):
        return 1

    used_parts = 1
    target_workbook = load_workbook(target_workbook_path)
    try:
        target_ws = target_workbook.active

        for source_path in source_paths[1:]:
            source_workbook = load_workbook(source_path)
            try:
                source_ws = source_workbook.active
                descriptive_count = descriptive_columns(target_ws, source_ws)
                if source_ws.max_column <= descriptive_count:
                    continue
                append_start_col = target_ws.max_column + 1
                copy_column_block(
                    target_ws,
                    source_ws,
                    descriptive_count + 1,
                    source_ws.max_column,
                    append_start_col,
                )
                used_parts += 1
            finally:
                source_workbook.close()

        target_workbook.save(target_workbook_path)
    finally:
        target_workbook.close()

    return used_parts


def cleanup_failed_group(source_paths: list[Path], source_dir: Path, target_dir: Path) -> None:
    target_workbook_path, target_json_path = target_paths_for_group(source_paths, source_dir, target_dir)

    if target_workbook_path.exists():
        target_workbook_path.unlink()
    if target_json_path.exists():
        target_json_path.unlink()


def append_debug_log(debug_path: Path, source_paths: list[Path], error_text: str) -> None:
    with debug_path.open("a", encoding="utf-8") as handle:
        handle.write("=" * 80 + "\n")
        handle.write(f"Failed group: {source_paths[0]}\n")
        handle.write("Source files:\n")
        for source_path in source_paths:
            handle.write(f"  {source_path}\n")
        handle.write("\n")
        handle.write(error_text.rstrip())
        handle.write("\n\n")


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(
    completed_jobs: int,
    total_jobs: int,
    merged_parts: int,
    failed_jobs: int,
    started_at: float,
) -> None:
    ratio = completed_jobs / total_jobs if total_jobs else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = completed_jobs / elapsed if elapsed > 0 else 0.0

    print(
        (
            f"\r[{bar}] {completed_jobs}/{total_jobs} files"
            f" | {merged_parts} parts"
            f" | {failed_jobs} errors"
            f" | {speed:5.1f} files/s"
            f" | {format_duration(elapsed)}"
        ),
        end="",
        flush=True,
    )


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()
    debug_path = (base_dir / DEBUG_FILE_NAME).resolve()

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    jobs = collect_jobs(source_dir)
    if not jobs:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    recreate_target_dir(target_dir)
    debug_path.write_text("", encoding="utf-8")

    print(f"Merging {len(jobs)} output files...", flush=True)

    started_at = time.monotonic()
    completed_jobs = 0
    merged_parts = 0
    failed_jobs = 0
    print_progress(completed_jobs, len(jobs), merged_parts, failed_jobs, started_at)

    for source_paths in jobs:
        try:
            merged_parts += merge_group(source_paths, source_dir, target_dir)
        except Exception:
            append_debug_log(debug_path, source_paths, traceback.format_exc())
            cleanup_failed_group(source_paths, source_dir, target_dir)
            failed_jobs += 1

        completed_jobs += 1
        print_progress(completed_jobs, len(jobs), merged_parts, failed_jobs, started_at)

    print()
    if failed_jobs:
        print(
            f"Merged {merged_parts} source part(s) into {completed_jobs - failed_jobs} output file(s) in {target_dir}",
            flush=True,
        )
        print(f"Logged {failed_jobs} error(s) to {debug_path}", flush=True)
        return 0

    print(f"Merged {merged_parts} source part(s) into {completed_jobs} output file(s) in {target_dir}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
