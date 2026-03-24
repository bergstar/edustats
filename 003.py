#!/usr/bin/env python3

from __future__ import annotations

import json
import os
import shutil
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange


SOURCE_DIR_NAME = "002_output"
TARGET_DIR_NAME = "003_output"
MAX_WORKERS = 8
BAR_WIDTH = 32


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def cell_has_content(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def column_is_merged(worksheet, column_index: int) -> bool:
    for cell_range in worksheet.merged_cells.ranges:
        if cell_range.min_col <= column_index <= cell_range.max_col:
            return True
    return False


def delete_trailing_empty_columns(worksheet) -> int:
    deleted_columns = 0

    while worksheet.max_column > 0:
        column_index = worksheet.max_column
        if column_is_merged(worksheet, column_index):
            break

        has_content = False
        for row in range(1, worksheet.max_row + 1):
            if cell_has_content(worksheet.cell(row=row, column=column_index).value):
                has_content = True
                break

        if has_content:
            break

        worksheet.delete_cols(column_index)
        deleted_columns += 1

    return deleted_columns


def effective_last_col(worksheet) -> int:
    max_value_col = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell_has_content(cell.value) and cell.column > max_value_col:
                max_value_col = cell.column

    max_merged_col = max((cell_range.max_col for cell_range in worksheet.merged_cells.ranges), default=0)
    return max(max_value_col, max_merged_col)


def extract_header(worksheet) -> tuple[int, dict[str, str], list[CellRange]]:
    last_col = effective_last_col(worksheet)
    original_merges = [CellRange(str(cell_range)) for cell_range in worksheet.merged_cells.ranges]
    header_values: dict[str, str] = {}
    header_rows = 0

    for row in range(1, min(worksheet.max_row, 10) + 1):
        full_width_merge = None
        for cell_range in original_merges:
            if (
                cell_range.min_row == row
                and cell_range.max_row == row
                and cell_range.min_col == 1
                and cell_range.max_col == last_col
            ):
                full_width_merge = cell_range
                break

        if full_width_merge is None:
            break

        value = worksheet.cell(row=row, column=1).value
        header_values[f"{row:02d}"] = "" if value is None else str(value)
        header_rows += 1

    if not 1 <= header_rows <= 10:
        raise ValueError(f"Expected 1 to 10 header rows, got {header_rows}")

    return header_rows, header_values, original_merges


def shifted_merges(
    original_merges: list[CellRange],
    deleted_rows: int,
) -> list[CellRange]:
    new_ranges: list[CellRange] = []
    for cell_range in original_merges:
        if cell_range.max_row <= deleted_rows:
            continue

        new_ranges.append(
            CellRange(
            min_col=cell_range.min_col,
            min_row=max(1, cell_range.min_row - deleted_rows),
            max_col=cell_range.max_col,
            max_row=cell_range.max_row - deleted_rows,
        )
        )

    return new_ranges


def process_file(source_path_str: str, source_dir_str: str, target_dir_str: str) -> int:
    source_path = Path(source_path_str)
    source_dir = Path(source_dir_str)
    target_dir = Path(target_dir_str)

    relative_path = source_path.relative_to(source_dir)
    target_workbook_path = target_dir / relative_path
    target_json_path = target_workbook_path.with_suffix(".json")

    target_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, target_workbook_path)

    workbook = load_workbook(target_workbook_path)
    try:
        worksheet = workbook.active
        delete_trailing_empty_columns(worksheet)
        header_rows, header_values, original_merges = extract_header(worksheet)
        new_merges = shifted_merges(original_merges, header_rows)
        worksheet.merged_cells.ranges = set()
        worksheet.delete_rows(1, header_rows)
        for cell_range in new_merges:
            worksheet.merge_cells(str(cell_range))
        workbook.save(target_workbook_path)
    finally:
        workbook.close()

    with target_json_path.open("w", encoding="utf-8") as json_file:
        json.dump(header_values, json_file, ensure_ascii=False, indent=2)
        json_file.write("\n")

    return header_rows


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(
    completed_files: int,
    total_files: int,
    header_lines: int,
    started_at: float,
) -> None:
    ratio = completed_files / total_files if total_files else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = completed_files / elapsed if elapsed > 0 else 0.0

    print(
        (
            f"\r[{bar}] {completed_files}/{total_files} files"
            f" | {header_lines} header lines"
            f" | {speed:5.1f} files/s"
            f" | {format_duration(elapsed)}"
        ),
        end="",
        flush=True,
    )


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    workbook_paths = iter_workbooks(source_dir)
    if not workbook_paths:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    recreate_target_dir(target_dir)

    worker_count = min(MAX_WORKERS, len(workbook_paths), os.cpu_count() or 1)
    print(
        f"Cleaning {len(workbook_paths)} files with {worker_count} workers...",
        flush=True,
    )

    started_at = time.monotonic()
    completed_files = 0
    header_lines = 0
    print_progress(completed_files, len(workbook_paths), header_lines, started_at)

    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        future_to_path = {
            executor.submit(
                process_file,
                str(workbook_path),
                str(source_dir),
                str(target_dir),
            ): workbook_path
            for workbook_path in workbook_paths
        }

        for future in as_completed(future_to_path):
            workbook_path = future_to_path[future]
            try:
                removed_header_rows = future.result()
            except Exception as exc:
                print()
                print(f"Failed while cleaning {workbook_path}: {exc}", file=sys.stderr)
                return 1

            completed_files += 1
            header_lines += removed_header_rows
            print_progress(
                completed_files,
                len(workbook_paths),
                header_lines,
                started_at,
            )

    print()
    print(
        f"Cleaned {completed_files} file(s) and stored {header_lines} header line(s) in {target_dir}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
