#!/usr/bin/env python3

from __future__ import annotations

import json
import shutil
import sys
import time
from copy import copy
from multiprocessing import cpu_count, get_context
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SOURCE_DIR_NAME = "004_output"
TARGET_DIR_NAME = "007_output"
REGIONS_LOOKUP_NAME = "regions_lookup.json"
BAR_WIDTH = 32


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def discover_regions(source_dir: Path) -> list[str]:
    return sorted({workbook_path.parent.name for workbook_path in iter_workbooks(source_dir)})


def build_default_regions_lookup(regions: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    used_codes: set[int] = set()

    if "г.москва" in regions:
        lookup["г.москва"] = "08"
        used_codes.add(8)

    next_code = 1
    for region in regions:
        if region in lookup:
            continue

        while next_code in used_codes:
            next_code += 1

        lookup[region] = f"{next_code:02d}"
        used_codes.add(next_code)
        next_code += 1

    return dict(sorted(lookup.items(), key=lambda item: item[1]))


def load_regions_lookup(lookup_path: Path, regions: list[str]) -> dict[str, str]:
    if lookup_path.exists():
        lookup = json.loads(lookup_path.read_text(encoding="utf-8"))
        if not isinstance(lookup, dict):
            raise ValueError(f"Invalid regions lookup in {lookup_path}")
    else:
        lookup = {}

    for region, code in list(lookup.items()):
        if not isinstance(region, str) or not isinstance(code, str):
            raise ValueError(f"Invalid region lookup entry: {region!r} -> {code!r}")
        lookup[region] = code.zfill(2)

    missing_regions = [region for region in regions if region not in lookup]
    if missing_regions:
        default_lookup = build_default_regions_lookup(regions)
        used_codes = {int(code) for code in lookup.values()}

        for region in missing_regions:
            preferred_code = int(default_lookup[region])
            if preferred_code not in used_codes:
                lookup[region] = f"{preferred_code:02d}"
                used_codes.add(preferred_code)
                continue

            next_code = 1
            while next_code in used_codes:
                next_code += 1
            lookup[region] = f"{next_code:02d}"
            used_codes.add(next_code)

    ordered_lookup = dict(sorted(lookup.items(), key=lambda item: item[1]))
    lookup_path.write_text(json.dumps(ordered_lookup, ensure_ascii=False, indent=2), encoding="utf-8")
    return ordered_lookup


def is_integer_like(value) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        return value.strip().isdigit()
    return False


def column_number_key(value) -> str:
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return value.strip()


def normalize_text(value) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text if text else None


def find_code_row(worksheet) -> int | None:
    scan_width = min(worksheet.max_column, 10)

    for row in range(1, worksheet.max_row + 1):
        values = [worksheet.cell(row=row, column=column).value for column in range(1, scan_width + 1)]
        if len(values) < 2:
            continue

        if not is_integer_like(values[0]) or int(float(values[0])) != 1:
            continue
        if not is_integer_like(values[1]) or int(float(values[1])) != 2:
            continue

        numeric_count = sum(1 for value in values if is_integer_like(value))
        if numeric_count >= 3:
            return row

    return None


def build_header_anchor_map(worksheet, code_row: int) -> dict[tuple[int, int], tuple[int, int]]:
    anchor_map: dict[tuple[int, int], tuple[int, int]] = {}

    for cell_range in worksheet.merged_cells.ranges:
        if cell_range.min_row >= code_row:
            continue

        row_end = min(cell_range.max_row, code_row - 1)
        for row in range(cell_range.min_row, row_end + 1):
            for column in range(cell_range.min_col, cell_range.max_col + 1):
                anchor_map[(row, column)] = (cell_range.min_row, cell_range.min_col)

    return anchor_map


def extract_library(worksheet) -> tuple[int, dict[str, dict[str, str]]]:
    code_row = find_code_row(worksheet)
    if code_row is None:
        raise ValueError("Could not find running column number row")

    anchor_map = build_header_anchor_map(worksheet, code_row)
    library: dict[str, dict[str, str]] = {}

    for column in range(1, worksheet.max_column + 1):
        code_value = worksheet.cell(row=code_row, column=column).value
        if not is_integer_like(code_value):
            continue

        chain: dict[str, str] = {}
        for row in range(code_row - 1, 0, -1):
            anchor_row, anchor_col = anchor_map.get((row, column), (row, column))
            text = normalize_text(worksheet.cell(row=anchor_row, column=anchor_col).value)
            if not text:
                continue
            chain.setdefault(str(anchor_row), text)

        library[column_number_key(code_value)] = chain

    return code_row, library


def find_header_block(worksheet, code_row: int) -> tuple[int, int] | None:
    header_rows: list[int] = []

    for row in range(1, code_row):
        for column in range(1, worksheet.max_column + 1):
            text = normalize_text(worksheet.cell(row=row, column=column).value)
            if text is None or is_integer_like(text):
                continue
            header_rows.append(row)
            break

    if not header_rows:
        return None

    return min(header_rows), max(header_rows)


def remove_text_headers(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook.active
        code_row = find_code_row(worksheet)
        if code_row is None:
            raise ValueError("Could not find running column number row")

        header_block = find_header_block(worksheet, code_row)
        if header_block is None:
            workbook.save(workbook_path)
            return

        start_row, end_row = header_block
        rows_to_remove = end_row - start_row + 1

        for cell_range in list(worksheet.merged_cells.ranges):
            intersects_header = not (
                cell_range.max_row < start_row or cell_range.min_row > end_row
            )
            if intersects_header:
                worksheet.unmerge_cells(str(cell_range))

        worksheet.delete_rows(start_row, rows_to_remove)
        workbook.save(workbook_path)
    finally:
        workbook.close()


def collect_jobs(source_dir: Path) -> list[list[Path]]:
    grouped_paths: dict[tuple[str, str, str], list[Path]] = {}

    for workbook_path in iter_workbooks(source_dir):
        relative_path = workbook_path.relative_to(source_dir)
        format_name, ownership_name, _, = relative_path.parts[:3]
        grouped_paths.setdefault((format_name, ownership_name, workbook_path.stem), []).append(workbook_path)

    jobs: list[list[Path]] = []
    for _, paths in sorted(grouped_paths.items()):
        paths.sort()
        jobs.append(paths)

    return jobs


def group_target_paths(
    source_paths: list[Path],
    source_dir: Path,
    target_dir: Path,
) -> tuple[Path, Path, Path]:
    first_path = sorted(source_paths)[0]
    relative_path = first_path.relative_to(source_dir)
    format_name, ownership_name = relative_path.parts[:2]
    target_workbook_path = target_dir / format_name / ownership_name / first_path.name
    target_json_path = target_workbook_path.with_suffix(".json")
    target_library_path = target_workbook_path.with_name(f"{target_workbook_path.stem}_library.json")
    return target_workbook_path, target_json_path, target_library_path


def region_name_for_path(workbook_path: Path, source_dir: Path) -> str:
    return workbook_path.relative_to(source_dir).parts[2]


def region_code_for_path(workbook_path: Path, source_dir: Path, regions_lookup: dict[str, str]) -> str:
    return regions_lookup[region_name_for_path(workbook_path, source_dir)]


def sort_group_paths(
    source_paths: list[Path],
    source_dir: Path,
    regions_lookup: dict[str, str],
) -> list[Path]:
    return sorted(
        source_paths,
        key=lambda path: (
            int(region_code_for_path(path, source_dir, regions_lookup)),
            region_name_for_path(path, source_dir),
        ),
    )


def running_number_sequence(worksheet, code_row: int, column_count: int) -> list[str | None]:
    sequence: list[str | None] = []

    for column in range(1, column_count + 1):
        value = worksheet.cell(row=code_row, column=column).value
        if is_integer_like(value):
            sequence.append(column_number_key(value))
            continue

        if normalize_text(value) is None:
            sequence.append(None)
            continue

        sequence.append(normalize_text(value))

    return sequence


def first_spravka_row(worksheet, start_row: int) -> int | None:
    for row in range(start_row, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column + 1):
            text = normalize_text(worksheet.cell(row=row, column=column).value)
            if text and text.lower().startswith("справка"):
                return row
    return None


def last_non_empty_row(worksheet, start_row: int = 1, end_row: int | None = None) -> int:
    if end_row is None:
        end_row = worksheet.max_row

    for row in range(end_row, start_row - 1, -1):
        for column in range(1, worksheet.max_column + 1):
            if normalize_text(worksheet.cell(row=row, column=column).value) is not None:
                return row
    return start_row - 1


def data_end_row(worksheet, code_row: int) -> int:
    start_row = code_row + 1
    spravka_row = first_spravka_row(worksheet, start_row)
    end_row = spravka_row - 1 if spravka_row is not None else worksheet.max_row
    return last_non_empty_row(worksheet, start_row, end_row)


def remove_spravka_block(worksheet, code_row: int) -> None:
    start_row = code_row + 1
    spravka_row = first_spravka_row(worksheet, start_row)
    if spravka_row is None:
        return
    worksheet.delete_rows(spravka_row, worksheet.max_row - spravka_row + 1)


def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def copy_row(source_ws, source_row: int, target_ws, target_row: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        copy_cell(
            source_ws.cell(row=source_row, column=column),
            target_ws.cell(row=target_row, column=column),
        )

    source_height = source_ws.row_dimensions[source_row].height
    if source_height is not None:
        target_ws.row_dimensions[target_row].height = source_height


def style_region_cell(worksheet, row: int, region_column: int) -> None:
    source_column = region_column - 1
    source_cell = worksheet.cell(row=row, column=source_column)
    target_cell = worksheet.cell(row=row, column=region_column)

    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)


def append_region_column(worksheet, code_row: int, region_code: str, data_end: int) -> tuple[int, str]:
    region_column = worksheet.max_column + 1
    existing_codes = [
        int(float(worksheet.cell(row=code_row, column=column).value))
        for column in range(1, worksheet.max_column + 1)
        if is_integer_like(worksheet.cell(row=code_row, column=column).value)
    ]
    next_code = str(max(existing_codes) + 1)

    worksheet.cell(row=code_row, column=region_column).value = int(next_code)
    style_region_cell(worksheet, code_row, region_column)

    previous_letter = get_column_letter(region_column - 1)
    region_letter = get_column_letter(region_column)
    previous_width = worksheet.column_dimensions[previous_letter].width
    worksheet.column_dimensions[region_letter].width = previous_width if previous_width is not None else 10

    if code_row > 1:
        worksheet.cell(row=1, column=region_column).value = "Регион"
        style_region_cell(worksheet, 1, region_column)
        worksheet.merge_cells(start_row=1, start_column=region_column, end_row=code_row - 1, end_column=region_column)

    for row in range(code_row + 1, data_end + 1):
        worksheet.cell(row=row, column=region_column).value = region_code
        style_region_cell(worksheet, row, region_column)

    return region_column, next_code


def merged_sidecar_data(first_path: Path) -> dict[str, str] | None:
    source_json_path = first_path.with_suffix(".json")
    if not source_json_path.exists():
        return None

    data = json.loads(source_json_path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        return None

    merged = {str(key): "" if value is None else str(value) for key, value in data.items()}
    merged["01"] = "Все регионы"
    return merged


def process_group(
    source_paths: list[Path],
    source_dir: Path,
    target_dir: Path,
    regions_lookup: dict[str, str],
) -> None:
    ordered_paths = sort_group_paths(source_paths, source_dir, regions_lookup)
    first_path = ordered_paths[0]
    target_workbook_path, target_json_path, target_library_path = group_target_paths(
        ordered_paths, source_dir, target_dir
    )

    target_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    target_workbook = load_workbook(first_path)
    try:
        target_ws = target_workbook.active
        code_row = find_code_row(target_ws)
        if code_row is None:
            raise ValueError(f"Could not find running column number row in {first_path}")

        source_column_count = target_ws.max_column
        header_sequence = running_number_sequence(target_ws, code_row, source_column_count)
        first_data_end = data_end_row(target_ws, code_row)
        region_column, region_code_key = append_region_column(
            target_ws,
            code_row,
            region_code_for_path(first_path, source_dir, regions_lookup),
            first_data_end,
        )
        remove_spravka_block(target_ws, code_row)
        append_row = last_non_empty_row(target_ws, 1, target_ws.max_row) + 1

        for source_path in ordered_paths[1:]:
            source_workbook = load_workbook(source_path)
            try:
                source_ws = source_workbook.active
                source_code_row = find_code_row(source_ws)
                if source_code_row is None:
                    raise ValueError(f"Could not find running column number row in {source_path}")

                source_header_sequence = running_number_sequence(source_ws, source_code_row, source_column_count)
                if source_header_sequence != header_sequence:
                    raise ValueError(f"Running column numbers do not match for {source_path}")

                source_data_end = data_end_row(source_ws, source_code_row)
                if source_data_end < source_code_row + 1:
                    continue

                source_region_code = region_code_for_path(source_path, source_dir, regions_lookup)

                for source_row in range(source_code_row + 1, source_data_end + 1):
                    copy_row(source_ws, source_row, target_ws, append_row, source_column_count)
                    target_ws.cell(row=append_row, column=region_column).value = source_region_code
                    style_region_cell(target_ws, append_row, region_column)
                    append_row += 1
            finally:
                source_workbook.close()

        _, library = extract_library(target_ws)
        library[region_code_key] = {"1": "Регион"}
        target_workbook.save(target_workbook_path)
    finally:
        target_workbook.close()

    target_library_path.write_text(json.dumps(library, ensure_ascii=False, indent=2), encoding="utf-8")

    merged_json = merged_sidecar_data(first_path)
    if merged_json is not None:
        target_json_path.write_text(json.dumps(merged_json, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_worker_count(argv: list[str]) -> int:
    if len(argv) < 2:
        return min(8, cpu_count() or 8)
    try:
        worker_count = int(argv[1])
    except ValueError as error:
        raise ValueError(f"Invalid worker count: {argv[1]}") from error
    if worker_count < 1:
        raise ValueError("Worker count must be at least 1")
    return worker_count


def process_group_job(args: tuple[list[str], str, str, dict[str, str]]) -> int:
    source_paths_raw, source_dir_raw, target_dir_raw, regions_lookup = args
    process_group(
        [Path(path) for path in source_paths_raw],
        Path(source_dir_raw),
        Path(target_dir_raw),
        regions_lookup,
    )
    return 1


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(completed_jobs: int, total_jobs: int, started_at: float) -> None:
    ratio = completed_jobs / total_jobs if total_jobs else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = completed_jobs / elapsed if elapsed > 0 else 0.0

    print(
        f"\r[{bar}] {completed_jobs}/{total_jobs} files | {speed:5.1f} files/s | {format_duration(elapsed)}",
        end="",
        flush=True,
    )


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()
    regions_lookup_path = (base_dir / REGIONS_LOOKUP_NAME).resolve()
    try:
        worker_count = parse_worker_count(sys.argv)
    except ValueError as error:
        print(str(error), file=sys.stderr)
        return 1

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    workbooks = iter_workbooks(source_dir)
    if not workbooks:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    jobs = collect_jobs(source_dir)
    recreate_target_dir(target_dir)
    regions_lookup = load_regions_lookup(regions_lookup_path, discover_regions(source_dir))

    print(f"Merging {len(jobs)} regional workbook groups with {worker_count} workers...", flush=True)
    started_at = time.monotonic()
    completed_jobs = 0
    print_progress(completed_jobs, len(jobs), started_at)

    if worker_count == 1:
        for source_paths in jobs:
            process_group(source_paths, source_dir, target_dir, regions_lookup)
            completed_jobs += 1
            print_progress(completed_jobs, len(jobs), started_at)
    else:
        job_args = [
            ([str(path) for path in source_paths], str(source_dir), str(target_dir), regions_lookup)
            for source_paths in jobs
        ]
        context_name = "fork" if sys.platform != "win32" else "spawn"
        with get_context(context_name).Pool(processes=worker_count, maxtasksperchild=5) as pool:
            for _ in pool.imap_unordered(process_group_job, job_args, chunksize=1):
                completed_jobs += 1
                print_progress(completed_jobs, len(jobs), started_at)

    print()
    print(f"Created merged cleaned workbooks and library JSON files in {target_dir}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
