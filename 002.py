#!/usr/bin/env python3

from __future__ import annotations

import argparse
import os
import queue
import shutil
import sys
import time
from concurrent.futures import FIRST_COMPLETED, ProcessPoolExecutor, wait
from multiprocessing import Manager
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR_NAME = "001_output"
TARGET_DIR_NAME = "002_output"
DEFAULT_WORKERS = 8
BAR_WIDTH = 32
POLL_INTERVAL = 0.05
SPINNER_FRAMES = "|/-\\"


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def collect_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def collect_workbook_plans(
    source_dir: Path, target_dir: Path, workbook_paths: list[Path]
) -> list[tuple[Path, Path, list[str]]]:
    plans: list[tuple[Path, Path, list[str]]] = []

    for workbook_path in workbook_paths:
        relative_region_dir = workbook_path.parent.relative_to(source_dir)
        destination_dir = target_dir / relative_region_dir
        sheet_names = collect_sheet_names(workbook_path)
        if sheet_names:
            plans.append((workbook_path, destination_dir, sheet_names))

    return plans


def export_workbook(
    workbook_path_str: str,
    destination_dir_str: str,
    sheet_names: list[str],
    progress_queue,
) -> int:
    workbook_path = Path(workbook_path_str)
    destination_dir = Path(destination_dir_str)
    exported_sheets = 0

    for sheet_name in sheet_names:
        workbook = load_workbook(workbook_path)
        try:
            for worksheet in list(workbook.worksheets):
                if worksheet.title != sheet_name:
                    workbook.remove(worksheet)

            workbook.active = 0
            destination_path = destination_dir / f"{sheet_name}.xlsx"
            destination_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(destination_path)
            exported_sheets += 1
            progress_queue.put(1)
        finally:
            workbook.close()

    return exported_sheets


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(
    frame_index: int,
    completed_books: int,
    total_books: int,
    exported_sheets: int,
    total_sheets: int,
    started_at: float,
) -> None:
    ratio = exported_sheets / total_sheets if total_sheets else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = exported_sheets / elapsed if elapsed > 0 else 0.0
    spinner = SPINNER_FRAMES[frame_index % len(SPINNER_FRAMES)]

    print(
        (
            f"\r{spinner} [{bar}] {exported_sheets}/{total_sheets} sheets"
            f" | {completed_books}/{total_books} books"
            f" | {speed:6.1f} sheets/s"
            f" | {format_duration(elapsed)}"
        ),
        end="",
        flush=True,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export workbook sheets into single-sheet files."
    )
    parser.add_argument(
        "workers",
        nargs="?",
        type=int,
        default=DEFAULT_WORKERS,
        help=f"Number of worker processes. Default: {DEFAULT_WORKERS}",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()

    if args.workers < 1:
        print("workers must be at least 1", file=sys.stderr)
        return 1

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    workbook_paths = iter_workbooks(source_dir)
    if not workbook_paths:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    workbook_plans = collect_workbook_plans(source_dir, target_dir, workbook_paths)
    if not workbook_plans:
        print(f"No sheets found in workbooks under {source_dir}", file=sys.stderr)
        return 1

    recreate_target_dir(target_dir)

    total_books = len(workbook_plans)
    total_sheets = sum(len(sheet_names) for _, _, sheet_names in workbook_plans)
    worker_count = min(args.workers, total_books, os.cpu_count() or 1)
    print(
        f"Exporting {total_books} workbooks | {total_sheets} sheets | {worker_count} workers",
        flush=True,
    )

    started_at = time.monotonic()
    completed_books = 0
    exported_sheets = 0
    frame_index = 0
    print_progress(
        frame_index,
        completed_books,
        total_books,
        exported_sheets,
        total_sheets,
        started_at,
    )

    with Manager() as manager:
        progress_queue = manager.Queue()

        with ProcessPoolExecutor(max_workers=worker_count) as executor:
            future_to_workbook = {}
            pending = set()

            for workbook_path, destination_dir, sheet_names in workbook_plans:
                future = executor.submit(
                    export_workbook,
                    str(workbook_path),
                    str(destination_dir),
                    sheet_names,
                    progress_queue,
                )
                future_to_workbook[future] = workbook_path
                pending.add(future)

            while pending:
                frame_index += 1

                while True:
                    try:
                        exported_sheets += progress_queue.get_nowait()
                    except queue.Empty:
                        break

                done, pending = wait(
                    pending,
                    timeout=POLL_INTERVAL,
                    return_when=FIRST_COMPLETED,
                )

                for future in done:
                    workbook_path = future_to_workbook[future]
                    try:
                        future.result()
                    except Exception as exc:
                        print()
                        print(
                            f"Failed while exporting {workbook_path}: {exc}",
                            file=sys.stderr,
                        )
                        return 1

                    completed_books += 1

                print_progress(
                    frame_index,
                    completed_books,
                    total_books,
                    exported_sheets,
                    total_sheets,
                    started_at,
                )

            while True:
                try:
                    exported_sheets += progress_queue.get_nowait()
                except queue.Empty:
                    break

            print_progress(
                frame_index,
                completed_books,
                total_books,
                exported_sheets,
                total_sheets,
                started_at,
            )

    print()
    print(
        f"Exported {exported_sheets} single-sheet file(s) from "
        f"{total_books} workbook(s) into {target_dir}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
