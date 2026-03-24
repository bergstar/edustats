#!/usr/bin/env python3

from __future__ import annotations

import json
import re
import shutil
import sys
import time
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR_NAME = "005_output"
TARGET_DIR_NAME = "006_output"
REGIONS_LOOKUP_NAME = "regions_lookup.json"
BAR_WIDTH = 32

FORMAT_CODES = {
    "full_time": "F",
    "hybrid": "H",
    "part_time": "P",
}

OWNERSHIP_CODES = {
    "commercial": "C",
    "governmental": "G",
}

DESCRIPTOR_LABEL_PATTERN = re.compile(r"[^а-яa-z0-9]+", re.IGNORECASE)


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def normalize_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    text = " ".join(str(value).split())
    return text if text else None


def normalized_label_key(value: str | None) -> str:
    if value is None:
        return ""
    return DESCRIPTOR_LABEL_PATTERN.sub("", value.lower())


def is_integer_like(value) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        return value.strip().isdigit()
    return False


def is_running_number_row(values: list[object], scan_width: int) -> bool:
    sample = values[:scan_width]
    if len(sample) < 2:
        return False

    if not is_integer_like(sample[0]) or int(float(sample[0])) != 1:
        return False
    if not is_integer_like(sample[1]) or int(float(sample[1])) != 2:
        return False

    numeric_count = sum(1 for value in sample if is_integer_like(value))
    return numeric_count >= 3


def sql_identifier(value: str) -> str:
    return f"`{value.replace('`', '``')}`"


def sql_literal(value) -> str:
    normalized = normalize_text(value)
    if normalized is None:
        return "NULL"
    escaped = normalized.replace("\\", "\\\\").replace("'", "''")
    return f"'{escaped}'"


def sql_rows(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    return ",\n".join(f"  ({', '.join(row)})" for row in rows)


def discover_regions(source_dir: Path) -> list[str]:
    return sorted({workbook_path.parent.name for workbook_path in iter_workbooks(source_dir)})


def build_default_regions_lookup(regions: list[str]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    used_codes: set[int] = set()

    if "г.москва" in regions:
        lookup["г.москва"] = "08"
        used_codes.add(8)

    next_code = 1
    for region in regions:
        if region in lookup:
            continue

        while next_code in used_codes:
            next_code += 1

        lookup[region] = f"{next_code:02d}"
        used_codes.add(next_code)
        next_code += 1

    return dict(sorted(lookup.items(), key=lambda item: item[1]))


def load_regions_lookup(lookup_path: Path, regions: list[str]) -> dict[str, str]:
    if lookup_path.exists():
        lookup = json.loads(lookup_path.read_text(encoding="utf-8"))
        if not isinstance(lookup, dict):
            raise ValueError(f"Invalid regions lookup in {lookup_path}")
    else:
        lookup = {}

    for region, code in list(lookup.items()):
        if not isinstance(region, str) or not isinstance(code, str):
            raise ValueError(f"Invalid region lookup entry: {region!r} -> {code!r}")
        lookup[region] = code.zfill(2)

    missing_regions = [region for region in regions if region not in lookup]
    if missing_regions:
        default_lookup = build_default_regions_lookup(regions)
        used_codes = {int(code) for code in lookup.values()}
        for region in missing_regions:
            preferred_code = int(default_lookup[region])
            if preferred_code not in used_codes:
                lookup[region] = f"{preferred_code:02d}"
                used_codes.add(preferred_code)
                continue

            next_code = 1
            while next_code in used_codes:
                next_code += 1
            lookup[region] = f"{next_code:02d}"
            used_codes.add(next_code)

        ordered_lookup = dict(sorted(lookup.items(), key=lambda item: item[1]))
        lookup_path.write_text(json.dumps(ordered_lookup, ensure_ascii=False, indent=2), encoding="utf-8")
        return ordered_lookup

    codes = list(lookup.values())
    if len(codes) != len(set(codes)):
        raise ValueError(f"Duplicate region codes in {lookup_path}")

    ordered_lookup = dict(sorted(lookup.items(), key=lambda item: item[1]))
    lookup_path.write_text(json.dumps(ordered_lookup, ensure_ascii=False, indent=2), encoding="utf-8")
    return ordered_lookup


def bundle_name(workbook_path: Path, source_dir: Path, regions_lookup: dict[str, str]) -> str:
    relative_path = workbook_path.relative_to(source_dir)
    format_name, ownership_name, region_name = relative_path.parts[:3]
    format_code = FORMAT_CODES[format_name]
    ownership_code = OWNERSHIP_CODES[ownership_name]
    region_code = regions_lookup[region_name]
    return f"{format_code}_{ownership_code}_{region_code}_{workbook_path.stem}"


def read_library(workbook_path: Path) -> dict[str, dict[str, str]]:
    library_path = workbook_path.with_name(f"{workbook_path.stem}_library.json")
    if not library_path.exists():
        raise ValueError(f"Missing library file: {library_path}")
    data = json.loads(library_path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"Invalid library JSON: {library_path}")
    return data


def read_cleaned_sheet(workbook_path: Path) -> tuple[int, list[int | None], list[tuple[int, list[object]]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        if worksheet.max_row < 1:
            return 1, [], []

        scan_width = min(worksheet.max_column, 10)
        number_row: int | None = None
        header_values: list[int | None] = []
        data_rows: list[tuple[int, list[object]]] = []

        for row_number, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(row_values)

            if number_row is None:
                if not is_running_number_row(values, scan_width):
                    continue

                number_row = row_number
                for column, value in enumerate(values, start=1):
                    if is_integer_like(value):
                        header_values.append(int(float(value)))
                        continue

                    if normalize_text(value) is None:
                        header_values.append(None)
                        continue

                    raise ValueError(f"Non-numeric header value in {workbook_path} at column {column}")
                continue

            data_rows.append((row_number, values))

        if number_row is None:
            raise ValueError(f"Could not find running number row in {workbook_path}")

        return number_row, header_values, data_rows
    finally:
        workbook.close()


def column_sql_name(column_number: int) -> str:
    return f"c{column_number:03d}"


def write_sql(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def primary_library_label(chain: dict[str, str]) -> str | None:
    if not chain:
        return None
    first_level = min(chain, key=lambda level: int(level))
    return chain[first_level]


def is_descriptor_label(label: str | None) -> bool:
    key = normalized_label_key(label)
    return key.startswith("наименование") or key.startswith("строки") or key.startswith("код")


def descriptor_prefix_count(library: dict[str, dict[str, str]]) -> int:
    count = 0
    code = 1

    while True:
        chain = library.get(str(code))
        if chain is None:
            break
        if not is_descriptor_label(primary_library_label(chain)):
            break
        count += 1
        code += 1

    return count


def infer_descriptor_column_count(
    header_values: list[int | None],
    library: dict[str, dict[str, str]],
    workbook_path: Path,
) -> int:
    prefix_count = descriptor_prefix_count(library)
    first_data_code = prefix_count + 1

    for position, header_value in enumerate(header_values, start=1):
        if header_value == first_data_code:
            return position - 1

    if prefix_count == 0:
        return 0

    raise ValueError(
        f"Could not infer descriptor columns in {workbook_path}: missing first data code {first_data_code}"
    )


def split_columns(
    header_values: list[int | None],
    descriptor_count: int,
    workbook_path: Path,
) -> tuple[list[int], list[int]]:
    descriptor_positions = list(range(descriptor_count))
    data_codes: list[int] = []

    for position in range(descriptor_count, len(header_values)):
        header_value = header_values[position]
        if header_value is None:
            raise ValueError(
                f"Missing data column number in {workbook_path} at physical column {position + 1}"
            )
        data_codes.append(header_value)

    return descriptor_positions, data_codes


def is_blank_workbook(data_rows: list[tuple[int, list[object]]]) -> bool:
    for _, row in data_rows:
        for value in row:
            if normalize_text(value) is not None:
                return False
    return True


def create_regions_sql(target_dir: Path, regions_lookup: dict[str, str]) -> None:
    table_name = "regions"
    rows = [
        [sql_literal(code), sql_literal(region)]
        for region, code in sorted(regions_lookup.items(), key=lambda item: item[1])
    ]
    insert_sql = sql_rows(rows)

    content = (
        f"DROP TABLE IF EXISTS {sql_identifier(table_name)};\n"
        f"CREATE TABLE {sql_identifier(table_name)} (\n"
        "  `region_code` VARCHAR(2) NOT NULL,\n"
        "  `region_name` LONGTEXT NOT NULL,\n"
        "  PRIMARY KEY (`region_code`)\n"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n"
    )
    if insert_sql:
        content += (
            f"INSERT INTO {sql_identifier(table_name)} (`region_code`, `region_name`) VALUES\n"
            f"{insert_sql};\n"
        )

    write_sql(target_dir / "regions.sql", content)


def create_column_dictionary_sql(
    target_dir: Path,
    table_name: str,
    data_codes: list[int],
    library: dict[str, dict[str, str]],
) -> None:
    max_level = 0
    for column_number in data_codes:
        chain = library.get(str(column_number), {})
        for level in chain:
            max_level = max(max_level, int(level))

    table_identifier = sql_identifier(f"{table_name}_column")
    level_columns = [f"  {sql_identifier(f'label_{level}')} LONGTEXT NULL" for level in range(1, max_level + 1)]
    create_columns = [
        "  `column_number` INT NOT NULL",
        "  `sql_column_name` VARCHAR(32) NOT NULL",
    ] + level_columns + [
        "  PRIMARY KEY (`column_number`)"
    ]

    rows: list[list[str]] = []
    for column_number in data_codes:
        chain = library.get(str(column_number), {})
        row = [
            str(column_number),
            sql_literal(column_sql_name(column_number)),
        ]
        for level in range(1, max_level + 1):
            row.append(sql_literal(chain.get(str(level))))
        rows.append(row)

    content = (
        f"DROP TABLE IF EXISTS {table_identifier};\n"
        f"CREATE TABLE {table_identifier} (\n"
        + ",\n".join(create_columns)
        + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n"
    )

    insert_sql = sql_rows(rows)
    if insert_sql:
        insert_columns = ["`column_number`", "`sql_column_name`"] + [
            sql_identifier(f"label_{level}") for level in range(1, max_level + 1)
        ]
        content += (
            f"INSERT INTO {table_identifier} ({', '.join(insert_columns)}) VALUES\n"
            f"{insert_sql};\n"
        )

    write_sql(target_dir / f"{table_name}_column.sql", content)


def create_row_dictionary_sql(
    target_dir: Path,
    table_name: str,
    descriptor_count: int,
    data_rows: list[tuple[int, list[object]]],
) -> None:
    table_identifier = sql_identifier(f"{table_name}_row")
    create_columns = [
        "  `row_id` INT NOT NULL",
        "  `sheet_row_number` INT NOT NULL",
    ] + [
        f"  {sql_identifier(f'c{index:03d}')} LONGTEXT NULL"
        for index in range(1, descriptor_count + 1)
    ] + [
        "  PRIMARY KEY (`row_id`)"
    ]

    rows: list[list[str]] = []
    for index, (sheet_row_number, row) in enumerate(data_rows, start=1):
        values = row[:descriptor_count]
        padded = values + [None] * (descriptor_count - len(values))
        rows.append(
            [str(index), str(sheet_row_number)]
            + [sql_literal(value) for value in padded[:descriptor_count]]
        )

    content = (
        f"DROP TABLE IF EXISTS {table_identifier};\n"
        f"CREATE TABLE {table_identifier} (\n"
        + ",\n".join(create_columns)
        + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n"
    )

    insert_sql = sql_rows(rows)
    if insert_sql:
        descriptor_columns = [sql_identifier(f"c{index:03d}") for index in range(1, descriptor_count + 1)]
        content += (
            f"INSERT INTO {table_identifier} (`row_id`, `sheet_row_number`"
            + (", " + ", ".join(descriptor_columns) if descriptor_columns else "")
            + ") VALUES\n"
            f"{insert_sql};\n"
        )

    write_sql(target_dir / f"{table_name}_row.sql", content)


def create_main_table_sql(
    target_dir: Path,
    table_name: str,
    data_codes: list[int],
    descriptor_count: int,
    data_rows: list[tuple[int, list[object]]],
) -> None:
    table_identifier = sql_identifier(table_name)

    create_columns = ["  `row_id` INT NOT NULL"]
    for column_number in data_codes:
        create_columns.append(f"  {sql_identifier(column_sql_name(column_number))} LONGTEXT NULL")
    create_columns.append("  PRIMARY KEY (`row_id`)")

    rows: list[list[str]] = []
    for index, (_, row) in enumerate(data_rows, start=1):
        values = row[descriptor_count:]
        padded = values + [None] * (len(data_codes) - len(values))
        rows.append([str(index)] + [sql_literal(value) for value in padded[: len(data_codes)]])

    content = (
        f"DROP TABLE IF EXISTS {table_identifier};\n"
        f"CREATE TABLE {table_identifier} (\n"
        + ",\n".join(create_columns)
        + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n"
    )

    insert_sql = sql_rows(rows)
    if insert_sql:
        insert_columns = ["`row_id`"] + [sql_identifier(column_sql_name(column_number)) for column_number in data_codes]
        content += (
            f"INSERT INTO {table_identifier} ({', '.join(insert_columns)}) VALUES\n"
            f"{insert_sql};\n"
        )

    write_sql(target_dir / f"{table_name}.sql", content)


def process_workbook(workbook_path: Path, source_dir: Path, target_dir: Path, regions_lookup: dict[str, str]) -> bool:
    table_name = bundle_name(workbook_path, source_dir, regions_lookup)
    _, header_values, data_rows = read_cleaned_sheet(workbook_path)
    if is_blank_workbook(data_rows):
        return False

    library = read_library(workbook_path)
    descriptor_count = infer_descriptor_column_count(header_values, library, workbook_path)
    _, data_codes = split_columns(header_values, descriptor_count, workbook_path)

    create_main_table_sql(target_dir, table_name, data_codes, descriptor_count, data_rows)
    create_column_dictionary_sql(target_dir, table_name, data_codes, library)
    create_row_dictionary_sql(target_dir, table_name, descriptor_count, data_rows)
    return True


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(completed_files: int, total_files: int, started_at: float) -> None:
    ratio = completed_files / total_files if total_files else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = completed_files / elapsed if elapsed > 0 else 0.0

    print(
        f"\r[{bar}] {completed_files}/{total_files} bundles | {speed:5.1f} bundles/s | {format_duration(elapsed)}",
        end="",
        flush=True,
    )


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()
    regions_lookup_path = (base_dir / REGIONS_LOOKUP_NAME).resolve()

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    workbooks = iter_workbooks(source_dir)
    if not workbooks:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    recreate_target_dir(target_dir)
    regions_lookup = load_regions_lookup(regions_lookup_path, discover_regions(source_dir))
    create_regions_sql(target_dir, regions_lookup)

    print(f"Building SQL bundles for {len(workbooks)} workbooks...", flush=True)
    started_at = time.monotonic()
    completed_files = 0
    created_files = 0
    skipped_files = 0
    print_progress(completed_files, len(workbooks), started_at)

    for workbook_path in workbooks:
        created = process_workbook(workbook_path, source_dir, target_dir, regions_lookup)
        if created:
            created_files += 1
        else:
            skipped_files += 1
        completed_files += 1
        print_progress(completed_files, len(workbooks), started_at)

    print()
    print(
        f"Created SQL files for {created_files} workbooks in {target_dir} and skipped {skipped_files} blank workbooks",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
