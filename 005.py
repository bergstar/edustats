#!/usr/bin/env python3

from __future__ import annotations

import json
import shutil
import sys
import time
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR_NAME = "004_output"
TARGET_DIR_NAME = "005_output"
BAR_WIDTH = 32


def should_skip(path: Path) -> bool:
    return path.name.startswith("~$") or path.name.startswith(".")


def iter_workbooks(source_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in source_dir.rglob("*.xlsx")
        if path.is_file() and not should_skip(path)
    )


def recreate_target_dir(target_dir: Path) -> None:
    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)


def is_integer_like(value) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        return value.strip().isdigit()
    return False


def column_number_key(value) -> str:
    if isinstance(value, float):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return value.strip()


def find_code_row(worksheet) -> int | None:
    scan_width = min(worksheet.max_column, 10)

    for row in range(1, worksheet.max_row + 1):
        values = [worksheet.cell(row=row, column=column).value for column in range(1, scan_width + 1)]
        if len(values) < 2:
            continue

        if not is_integer_like(values[0]) or int(float(values[0])) != 1:
            continue
        if not is_integer_like(values[1]) or int(float(values[1])) != 2:
            continue

        numeric_count = sum(1 for value in values if is_integer_like(value))
        if numeric_count >= 3:
            return row

    return None


def normalize_text(value) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text if text else None


def build_header_anchor_map(worksheet, code_row: int) -> dict[tuple[int, int], tuple[int, int]]:
    anchor_map: dict[tuple[int, int], tuple[int, int]] = {}

    for cell_range in worksheet.merged_cells.ranges:
        if cell_range.min_row >= code_row:
            continue

        row_end = min(cell_range.max_row, code_row - 1)
        for row in range(cell_range.min_row, row_end + 1):
            for column in range(cell_range.min_col, cell_range.max_col + 1):
                anchor_map[(row, column)] = (cell_range.min_row, cell_range.min_col)

    return anchor_map


def extract_library(worksheet) -> tuple[int, dict[str, dict[str, str]]]:
    code_row = find_code_row(worksheet)
    if code_row is None:
        raise ValueError("Could not find running column number row")

    anchor_map = build_header_anchor_map(worksheet, code_row)
    library: dict[str, dict[str, str]] = {}

    for column in range(1, worksheet.max_column + 1):
        code_value = worksheet.cell(row=code_row, column=column).value
        if not is_integer_like(code_value):
            continue

        chain: dict[str, str] = {}
        for row in range(code_row - 1, 0, -1):
            anchor_row, anchor_col = anchor_map.get((row, column), (row, column))
            text = normalize_text(worksheet.cell(row=anchor_row, column=anchor_col).value)
            if not text:
                continue
            chain.setdefault(str(anchor_row), text)

        library[column_number_key(code_value)] = chain

    return code_row, library


def find_header_block(worksheet, code_row: int) -> tuple[int, int] | None:
    header_rows: list[int] = []

    for row in range(1, code_row):
        for column in range(1, worksheet.max_column + 1):
            text = normalize_text(worksheet.cell(row=row, column=column).value)
            if text is None or is_integer_like(text):
                continue
            header_rows.append(row)
            break

    if not header_rows:
        return None

    return min(header_rows), max(header_rows)


def remove_text_headers(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook.active
        code_row = find_code_row(worksheet)
        if code_row is None:
            raise ValueError("Could not find running column number row")

        header_block = find_header_block(worksheet, code_row)
        if header_block is None:
            workbook.save(workbook_path)
            return

        start_row, end_row = header_block
        rows_to_remove = end_row - start_row + 1

        for cell_range in list(worksheet.merged_cells.ranges):
            intersects_header = not (
                cell_range.max_row < start_row or cell_range.min_row > end_row
            )
            if intersects_header:
                worksheet.unmerge_cells(str(cell_range))

        worksheet.delete_rows(start_row, rows_to_remove)
        workbook.save(workbook_path)
    finally:
        workbook.close()


def copy_sidecar_json(source_workbook_path: Path, target_workbook_path: Path) -> None:
    source_json_path = source_workbook_path.with_suffix(".json")
    if source_json_path.exists():
        shutil.copy2(source_json_path, target_workbook_path.with_suffix(".json"))


def format_duration(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes, seconds = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def print_progress(completed_files: int, total_files: int, started_at: float) -> None:
    ratio = completed_files / total_files if total_files else 1.0
    filled = int(BAR_WIDTH * ratio)
    bar = "#" * filled + "-" * (BAR_WIDTH - filled)
    elapsed = time.monotonic() - started_at
    speed = completed_files / elapsed if elapsed > 0 else 0.0

    print(
        f"\r[{bar}] {completed_files}/{total_files} files | {speed:5.1f} files/s | {format_duration(elapsed)}",
        end="",
        flush=True,
    )


def process_workbook(source_workbook_path: Path, source_dir: Path, target_dir: Path) -> None:
    relative_path = source_workbook_path.relative_to(source_dir)
    target_workbook_path = target_dir / relative_path
    target_library_path = target_workbook_path.with_name(f"{target_workbook_path.stem}_library.json")

    target_workbook_path.parent.mkdir(parents=True, exist_ok=True)

    source_workbook = load_workbook(source_workbook_path, data_only=False)
    try:
        _, library = extract_library(source_workbook.active)
    finally:
        source_workbook.close()

    shutil.copy2(source_workbook_path, target_workbook_path)
    copy_sidecar_json(source_workbook_path, target_workbook_path)
    remove_text_headers(target_workbook_path)
    target_library_path.write_text(json.dumps(library, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    source_dir = (base_dir / SOURCE_DIR_NAME).resolve()
    target_dir = (base_dir / TARGET_DIR_NAME).resolve()

    if not source_dir.is_dir():
        print(f"Source directory does not exist: {source_dir}", file=sys.stderr)
        return 1

    workbooks = iter_workbooks(source_dir)
    if not workbooks:
        print(f"No .xlsx files found in {source_dir}", file=sys.stderr)
        return 1

    recreate_target_dir(target_dir)

    print(f"Converting {len(workbooks)} workbooks...", flush=True)
    started_at = time.monotonic()
    completed_files = 0
    print_progress(completed_files, len(workbooks), started_at)

    for workbook_path in workbooks:
        process_workbook(workbook_path, source_dir, target_dir)
        completed_files += 1
        print_progress(completed_files, len(workbooks), started_at)

    print()
    print(f"Created cleaned workbooks and library JSON files in {target_dir}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
